"""
Módulo de Exportação - biblioteca-br

Funções para exportar DataFrames para Excel, CSV, PDF e JSON.

Autor: Isaque Sena
"""

import pandas as pd
import os
from datetime import datetime
from typing import Optional, Union, List


def excel(
    df: pd.DataFrame,
    caminho: str,
    nome_planilha: str = "Dados",
    formato_data: str = "dd/mm/yyyy",
    adicionar_formatacao: bool = True,
    autor: str = "biblioteca-br"
) -> str:
    """
    Exporta DataFrame para Excel com formatação profissional.
    
    Args:
        df: DataFrame pandas a ser exportado
        caminho: Caminho do arquivo de saída (.xlsx)
        nome_planilha: Nome da planilha dentro do Excel
        formato_data: Formato das datas (padrão: dd/mm/yyyy)
        adicionar_formatacao: Se True, aplica formatação condicional e estilos
        autor: Nome do autor/metadados
    
    Returns:
        Caminho do arquivo criado
    
    Exemplo:
        >>> from biblioteca_br import bacen, exportar
        >>> df = bacen.selic()
        >>> exportar.excel(df, "selic_2024.xlsx")
    """
    
    # Garantir extensão .xlsx
    if not caminho.endswith('.xlsx'):
        caminho += '.xlsx'
    
    # Criar diretório se não existir
    diretorio = os.path.dirname(caminho)
    if diretorio and not os.path.exists(diretorio):
        os.makedirs(diretorio)
    
    # Criar writer com engine openpyxl
    with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=nome_planilha, index=False)
        
        if adicionar_formatacao:
            _formatar_excel(writer, df, nome_planilha, formato_data, autor)
    
    return caminho


def _formatar_excel(writer, df: pd.DataFrame, nome_planilha: str, formato_data: str, autor: str):
    """Aplica formatação profissional ao Excel."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    import openpyxl
    
    workbook = writer.book
    worksheet = writer.sheets[nome_planilha]
    
    # Cores e estilos
    cor_primaria = "4472C4"  # Azul profissional
    cor_cabecalho = "D9E1F2"  # Azul claro
    cor_alternada = "F8F9FA"  # Cinza muito claro
    fonte_cabecalho = Font(bold=True, color="FFFFFF", size=11)
    fonte_normal = Font(size=10)
    alinhamento_centro = Alignment(horizontal='center', vertical='center')
    alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
    
    # Borda fina
    borda_fina = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Preenchimento do cabeçalho
    fill_cabecalho = PatternFill(start_color=cor_primaria, end_color=cor_primaria, fill_type="solid")
    fill_alternado = PatternFill(start_color=cor_alternada, end_color=cor_alternada, fill_type="solid")
    
    # Formatar cabeçalho
    for col_num, coluna in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = str(coluna).upper()
        cell.font = fonte_cabecalho
        cell.fill = fill_cabecalho
        cell.alignment = alinhamento_centro
        cell.border = borda_fina
    
    # Formatar dados
    for row_num, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        fill_row = fill_alternado if row_num % 2 == 0 else None
        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = fonte_normal
            cell.alignment = alinhamento_esquerda
            cell.border = borda_fina
            
            # Formatar números como moeda ou porcentagem se apropriado
            if isinstance(value, (int, float)):
                if abs(value) < 1 and value != 0:
                    cell.number_format = '0.00%'
                elif abs(value) > 1000:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '0.00'
            
            # Formatar data
            elif isinstance(value, (datetime, pd.Timestamp)):
                cell.number_format = formato_data
            
            if fill_row:
                cell.fill = fill_row
    
    # Ajustar largura das colunas
    for col_num, coluna in enumerate(df.columns, 1):
        max_length = max(
            len(str(coluna)),
            df[coluna].astype(str).map(len).max() if len(df) > 0 else 0
        ) + 2
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = min(max_length, 50)
    
    # Congelar painéis (cabeçalho fixo)
    worksheet.freeze_panes = 'A2'
    
    # Adicionar metadados
    workbook.properties.title = nome_planilha
    workbook.properties.creator = autor
    workbook.properties.created = datetime.now()


def csv(
    df: pd.DataFrame,
    caminho: str,
    separador: str = ',',
    codificacao: str = 'utf-8',
    indice: bool = False,
    decimais: str = '.'
) -> str:
    """
    Exporta DataFrame para CSV.
    
    Args:
        df: DataFrame pandas a ser exportado
        caminho: Caminho do arquivo de saída (.csv)
        separador: Separador de campos (padrão: vírgula)
        codificacao: Codificação do arquivo (padrão: utf-8)
        indice: Se True, inclui índice no CSV
        decimais: Separador decimal (padrão: ponto)
    
    Returns:
        Caminho do arquivo criado
    
    Exemplo:
        >>> exportar.csv(df, "dados.csv", separador=';')  # CSV brasileiro
    """
    if not caminho.endswith('.csv'):
        caminho += '.csv'
    
    # Criar diretório se não existir
    diretorio = os.path.dirname(caminho)
    if diretorio and not os.path.exists(diretorio):
        os.makedirs(diretorio)
    
    df.to_csv(
        caminho,
        sep=separador,
        encoding=codificacao,
        index=indice,
        decimal=decimais,
        date_format='%Y-%m-%d'
    )
    
    return caminho


def pdf(
    df: pd.DataFrame,
    caminho: str,
    titulo: str = "Relatório de Dados",
    subtitulo: str = "",
    incluir_grafico: bool = False,
    grafico_df: Optional[pd.DataFrame] = None,
    colunas_grafico: Optional[List[str]] = None,
    autor: str = "biblioteca-br"
) -> str:
    """
    Gera relatório em PDF com tabela de dados e gráfico opcional.
    
    Args:
        df: DataFrame com os dados principais
        caminho: Caminho do arquivo PDF de saída
        titulo: Título do relatório
        subtitulo: Subtítulo opcional
        incluir_grafico: Se True, inclui um gráfico no PDF
        grafico_df: DataFrame específico para o gráfico (se None, usa df)
        colunas_grafico: Colunas para plotar no gráfico
        autor: Autor do relatório
    
    Returns:
        Caminho do arquivo criado
    
    Exemplo:
        >>> df = bacen.selic()
        >>> exportar.pdf(df, "relatorio_selic.pdf", titulo="Taxa Selic 2024")
    """
    if not caminho.endswith('.pdf'):
        caminho += '.pdf'
    
    # Criar diretório se não existir
    diretorio = os.path.dirname(caminho)
    if diretorio and not os.path.exists(diretorio):
        os.makedirs(diretorio)
    
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import matplotlib.pyplot as plt
    import io
    
    # Configurar documento
    doc = SimpleDocTemplate(
        caminho,
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.7*inch,
        bottomMargin=0.5*inch
    )
    
    elementos = []
    estilos = getSampleStyleSheet()
    
    # Estilo personalizado para título
    estilo_titulo = ParagraphStyle(
        'TituloPersonalizado',
        parent=estilos['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#4472C4'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    estilo_subtitulo = ParagraphStyle(
        'SubtituloPersonalizado',
        parent=estilos['Normal'],
        fontSize=12,
        textColor=colors.grey,
        spaceAfter=20,
        alignment=TA_CENTER,
        fontStyle='italic'
    )
    
    # Adicionar título
    elementos.append(Paragraph(titulo, estilo_titulo))
    
    if subtitulo:
        elementos.append(Paragraph(subtitulo, estilo_subtitulo))
    else:
        elementos.append(Spacer(1, 0.2*inch))
    
    # Adicionar informações do relatório
    info_texto = f"""
    <b>Gerado por:</b> {autor}<br/>
    <b>Data:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>
    <b>Registros:</b> {len(df)} | <b>Colunas:</b> {len(df.columns)}
    """
    estilo_info = ParagraphStyle(
        'Info',
        parent=estilos['Normal'],
        fontSize=9,
        textColor=colors.darkgrey,
        spaceAfter=15
    )
    elementos.append(Paragraph(info_texto, estilo_info))
    elementos.append(Spacer(1, 0.2*inch))
    
    # Preparar tabela
    dados_tabela = [df.columns.tolist()] + df.head(50).values.tolist()  # Máximo 50 linhas
    
    if len(df) > 50:
        dados_tabela.append(['...'])
        nota = f"* Mostrando 50 de {len(df)} registros. Exporte para Excel para ver todos."
        elementos.append(Paragraph(f"<i>{nota}</i>", ParagraphStyle('Nota', parent=estilos['Normal'], fontSize=8)))
        elementos.append(Spacer(1, 0.1*inch))
    
    tabela = Table(dados_tabela, repeatRows=1)
    
    # Estilizar tabela
    estilo_tabela = TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Linhas de dados
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ])
    
    tabela.setStyle(estilo_tabela)
    elementos.append(tabela)
    
    # Adicionar gráfico se solicitado
    if incluir_grafico:
        elementos.append(Spacer(1, 0.3*inch))
        elementos.append(Paragraph("<b>Visualização Gráfica</b>", estilos['Heading2']))
        elementos.append(Spacer(1, 0.1*inch))
        
        # Criar gráfico
        df_plot = grafico_df if grafico_df is not None else df
        
        fig, ax = plt.subplots(figsize=(10, 5), dpi=150)
        
        if colunas_grafico:
            cols_para_plotar = colunas_grafico
        else:
            cols_numericas = df_plot.select_dtypes(include=['number']).columns
            cols_para_plotar = cols_numericas[:3].tolist() if len(cols_numericas) > 0 else []
        
        if cols_para_plotar and len(df_plot) > 0:
            for col in cols_para_plotar:
                if col in df_plot.columns:
                    ax.plot(df_plot.index, df_plot[col], label=col, linewidth=2)
            
            ax.legend(loc='best')
            ax.grid(True, alpha=0.3)
            ax.set_xlabel('Período')
            ax.set_ylabel('Valor')
            plt.xticks(rotation=45)
            plt.tight_layout()
        else:
            ax.text(0.5, 0.5, 'Sem dados numéricos para plotar', 
                   transform=ax.transAxes, ha='center', va='center')
        
        # Salvar gráfico em buffer
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', bbox_inches='tight')
        buffer.seek(0)
        plt.close()
        
        # Adicionar imagem ao PDF
        img = Image(buffer, width=6*inch, height=3*inch)
        elementos.append(img)
    
    # Rodapé
    elementos.append(Spacer(1, 0.3*inch))
    rodape = Paragraph(
        f"<i>Documento gerado automaticamente por biblioteca-br • {datetime.now().strftime('%Y')}</i>",
        ParagraphStyle('Rodape', parent=estilos['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    )
    elementos.append(rodape)
    
    # Construir PDF
    doc.build(elementos)
    
    return caminho


def json(
    df: pd.DataFrame,
    caminho: str,
    orient: str = 'records',
    indent: int = 2,
    codificacao: str = 'utf-8'
) -> str:
    """
    Exporta DataFrame para JSON.
    
    Args:
        df: DataFrame pandas a ser exportado
        caminho: Caminho do arquivo de saída (.json)
        orient: Orientação do JSON ('records', 'list', 'split', 'index', 'columns')
        indent: Indentação para formatação (padrão: 2)
        codificacao: Codificação do arquivo
    
    Returns:
        Caminho do arquivo criado
    
    Exemplo:
        >>> exportar.json(df, "dados.json", orient='records')
    """
    if not caminho.endswith('.json'):
        caminho += '.json'
    
    # Criar diretório se não existir
    diretorio = os.path.dirname(caminho)
    if diretorio and not os.path.exists(diretorio):
        os.makedirs(diretorio)
    
    df.to_json(
        caminho,
        orient=orient,
        indent=indent,
        force_ascii=False,
        date_format='iso'
    )
    
    return caminho


# Alias para compatibilidade
to_excel = excel
to_csv = csv
to_pdf = pdf
to_json = json